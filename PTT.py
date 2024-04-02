import tkinter as tk
from tkinter import messagebox, ttk
import os
import datetime
import openpyxl
import psutil
import threading
import subprocess
from PIL import Image, ImageTk

# global var
is_sw_running = False
is_gcad_running = False
pause_sw_time = None
pause_gcad_time = None

class swTimeTracker:

    def __init__(self, master):
        self.logo_sw = None
        self.logo_gcad = None
        self.proj_num = None
        self.cl_name = None
        self.dwg_type = None
        self.designer_name = None
        self.remark = None
        self.start_time = None
        self.current_date = None
        self.current_sw_time = None
        self.current_gcad_time = None
        self.sw_path_chk = None
        self.gcad_path_chk = None
        self.one_drive_path = None
        self.master = master
        master.title("Project Time Tracker")

        master.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.img_sw = Image.open("SolidWorks.png")
        self.resize_logo_sw = self.img_sw.resize((50, 50))
        self.logo_sw = ImageTk.PhotoImage(self.resize_logo_sw)

        self.img_gcad = Image.open("GStarCAD.png")
        self.resize_logo_gcad = self.img_gcad.resize((50, 50))
        self.logo_gcad = ImageTk.PhotoImage(self.resize_logo_gcad)

        master.columnconfigure(0, weight=2)
        master.columnconfigure(1, weight=2)
        master.columnconfigure(2, weight=2)

        self.projnum_label = tk.Label(master, text="Project Number :", font=("Century Gothic Bold", 10), bg='linen')
        self.projnum_label.grid(row=0, column=0, sticky='w', padx=20, pady=5)

        self.projnum_entry = tk.Entry(master, fg='gray')
        self.projnum_entry.insert(0, "XX-XXXX")
        self.projnum_entry.grid(row=0, column=1, ipadx=30, sticky='w')

        self.clname_label = tk.Label(master, text="Client Name :", font=("Century Gothic Bold", 10), bg='linen')
        self.clname_label.grid(row=1, column=0, sticky='w', padx=20, pady=5)

        self.clname_entry = tk.Entry(master, fg='gray')
        self.clname_entry.insert(0, "Enter client name")
        self.clname_entry.grid(row=1, column=1, ipadx=30, sticky='w')

        self.dwgtype_label = tk.Label(master, text="Drawing Type :", font=("Century Gothic Bold", 10), bg='linen')
        self.dwgtype_label.grid(row=2, column=0, sticky='w', padx=20, pady=5)

        self.dwgtype_values = ['GA', 'FAB', 'CP', 'AsB']
        self.dwg_type_cb = ttk.Combobox(master, values=self.dwgtype_values, width=17, state='readonly')
        self.dwg_type_cb.current(0)
        self.dwg_type_cb.grid(row=2, column=1, ipadx=30, sticky='w')

        self.note_label = tk.Label(master, text="Note/Remarks :", font=("Century Gothic Bold", 10), bg='linen')
        self.note_label.place(x=20, y=100)

        self.note_text = tk.Text(master, font=("Segoe UI", 10))
        self.note_text.place(x=20, y=125, height=40, width=325)

        self.attn_label = tk.Label(master, text="!! PLEASE WAIT BUTTON COLOR TO CHANGE BEFORE CLOSING !!", font=("Century Gothic Bold", 10), bg='linen', foreground='#dd2c00', justify='center', wraplength=210)
        self.attn_label.place(relx=0.5, y=200, anchor='center')

        self.add_on_click(self.projnum_entry)
        self.add_on_click(self.clname_entry)
        
        self.logo_sw_btn = tk.Button(master, image=self.logo_sw, font=("Century Gothic Bold", 12), background='white', activebackground='limegreen', disabledforeground='darkgrey', command=self.start_sw)
        self.logo_sw_btn.place(x=90, y=260, anchor='center')

        self.logo_gcad_btn = tk.Button(master, image=self.logo_gcad, font=("Century Gothic Bold", 12), background='white', activebackground='limegreen', disabledforeground='darkgrey', command=self.start_gcad)
        self.logo_gcad_btn.place(x=270, y=260, anchor='center')
        
        self.timer_sw_label = tk.Label(master, font=("Century Gothic Bold", 12), bg='white', width=10, height=2, borderwidth='1', relief='sunken', justify='center')
        self.timer_sw_label.place(x=90, y=320, anchor='center')

        self.timer_gcad_label = tk.Label(master, font=("Century Gothic Bold", 12), bg='white', width=10, height=2, borderwidth='1', relief='sunken', justify='center')
        self.timer_gcad_label.place(x=270, y=320, anchor='center')

        self.pause_sw_btn = tk.Button(master, text='Pause', font=("Century Gothic Bold", 10), bg='#dd2c00', fg='white',width=9, height=1, relief='ridge', command=self.pause_sw_timer)
        self.pause_sw_btn.place(x=90, y=365, anchor='center')

        self.pause_gcad_btn = tk.Button(master, text='Pause', font=("Century Gothic Bold", 10), bg='#205bfa', fg='white',width=9, height=1, relief='ridge', command=self.pause_gcad_timer)
        self.pause_gcad_btn.place(x=270, y=365, anchor='center')
        
    # timer for SolidWorks
    def timer_sw_start(self):
        global is_sw_running, start_sw_time, pause_sw_time
        if not is_sw_running:
            if pause_sw_time is None:
                start_sw_time = datetime.datetime.now()
            else:
                start_sw_time += (datetime.datetime.now() - pause_sw_time)
                pause_sw_time = None
            
            is_sw_running = True
            self.update_sw_time()

        else:
            is_sw_running = False
            pause_sw_time = datetime.datetime.now()

    def timer_sw_stop(self):
        global is_sw_running
        is_sw_running = False

    def update_sw_time(self):
        if is_sw_running:
            stop_sw_time = datetime.datetime.now()
            elapsed_sw_time = stop_sw_time - start_sw_time
            elapsed_sw_time_str = str(elapsed_sw_time).split('.')[0]
            self.timer_sw_label.config(text=elapsed_sw_time_str)
            self.timer_sw_label.after(1000, self.update_sw_time)

    # timer for GstarCAD
    def timer_gcad_start(self):
        global is_gcad_running, start_gcad_time, pause_gcad_time
        if not is_gcad_running:
            if pause_gcad_time is None:
                start_gcad_time = datetime.datetime.now()
            else:
                start_gcad_time += (datetime.datetime.now() - pause_gcad_time)
                pause_gcad_time = None

            is_gcad_running = True
            self.update_gcad_time()

    def timer_gcad_stop(self):
        global is_gcad_running
        is_gcad_running = False

    def update_gcad_time(self):
        if is_gcad_running:
            stop_gcad_time = datetime.datetime.now()
            elapsed_gcad_time = stop_gcad_time - start_gcad_time
            elapsed_gcad_time_str = str(elapsed_gcad_time).split('.')[0]
            self.timer_gcad_label.config(text=elapsed_gcad_time_str)
            self.timer_gcad_label.after(1000, self.update_gcad_time)

    def pause_sw_timer(self):
        global is_sw_running, pause_sw_time
        is_sw_running = False
        pause_sw_time = datetime.datetime.now()
        self.pause_sw_btn.config(text='Resume', command=self.resume_sw_timer)

    def resume_sw_timer(self):
        global is_sw_running, start_sw_time, pause_sw_time
        if not is_sw_running and pause_sw_time is not None:
            start_sw_time += (datetime.datetime.now() - pause_sw_time)
            pause_sw_time = None
            is_sw_running = True
            self.pause_sw_btn.config(text='Rehat Jap', command=self.pause_sw_timer)
            self.update_sw_time()

    def pause_gcad_timer(self):
        global is_gcad_running, pause_gcad_time
        is_gcad_running = False
        pause_gcad_time = datetime.datetime.now()
        self.pause_gcad_btn.config(text='Resume', command=self.resume_gcad_timer)

    def resume_gcad_timer(self):
        global is_gcad_running, start_gcad_time, pause_gcad_time
        if not is_gcad_running and pause_gcad_time is not None:
            start_gcad_time += (datetime.datetime.now() - pause_gcad_time)
            pause_gcad_time = None
            is_gcad_running = True
            self.pause_gcad_btn.config(text='Rehat Jap', command=self.pause_gcad_timer)
            self.update_gcad_time()

    #when user click entry box, auto clear the placeholder
    def add_on_click(self, entry_widget):
        entry_widget.bind("<FocusIn>", lambda event: self.clear_entry(event.widget))

    def clear_entry(self, entry_widget):
        placeholder_texts = ["XX-XXXX", "Enter client name", "CPT/GA/FAB/FLT"]
        if entry_widget.get() in placeholder_texts:
            entry_widget.delete(0, "end")
            entry_widget.config(fg='black')

    # start SolidWorks multi-threading
    def start_sw(self):
        self.logo_sw_btn.config(state='disabled')

        t = threading.Thread(target=self.start_sw_thread)
        t.start()

    # get User input, start SolidWorks & Timer
    def start_sw_thread(self):
        self.proj_num = self.projnum_entry.get()
        self.cl_name = self.clname_entry.get().title()
        self.dwg_type = self.dwg_type = self.dwg_type_cb.get()
        self.remark = self.note_text.get("1.0", 'end').capitalize()

        self.current_date = datetime.date.today()
        self.current_sw_time = datetime.datetime.now()

        # check SW user path that exists
        self.sw_path_chk = os.path.exists(r"C:\Program Files\SOLIDWORKS Corp\SOLIDWORKS\SLDWORKS.exe")

        # for different user devices path
        if self.sw_path_chk == True:
            subprocess.Popen(r'"C:\Program Files\SOLIDWORKS Corp\SOLIDWORKS\SLDWORKS.exe"')

        else:
            subprocess.Popen(r'"C:\Program Files\SOLIDWORKS Corp 2021\SOLIDWORKS\SLDWORKS.exe"')

        self.timer_sw_start()

        # check applications in the background
        while True:
            if any("SLDWORKS.exe" in p.name() for p in psutil.process_iter()):
                # Solidworks is running, continue loop
                continue

            else:
                self.stop_sw()

                break

    # stop SolidWorks, open excel and save all the info
    def stop_sw(self):

        self.one_drive_path = os.environ.get('OneDriveCommercial') # get  OneDrive path
        xl_path_chk = [
            self.one_drive_path + r'\xxx\xxx\xxx\Project-Time-Table-2024.xlsx',
            self.one_drive_path + r'\xxx\xxx\xxx\xxx\Project-Time-Table-2024.xlsx',
            self.one_drive_path + r'\xxx\xxx\xxx\xxx\Project-Time-Table-2024.xlsx'
        ]

        self.designer = os.getlogin()
        self.designer_list = {
            'Device Name 1': 'Name 1',
            'Device Name 2': 'Name 2',
            'Device Name 3': 'Name 3',
            'Device Name 4': 'Name 4'
        }

        self.timer_sw_stop()
        self.logo_sw_btn.config(state='normal')

        stop_sw_time = datetime.datetime.now()
        elapsed_sw_time = stop_sw_time - start_sw_time

        # check where user excel path is
        for path in xl_path_chk:
            xlp = os.path.exists(path)

            if xlp == True:
                xl_path = path

        # check which user is signing in
        if self.designer in self.designer_list:
            self.designer_name = self.designer_list[self.designer]

        # open excel, insert all info, save then exit excel
        wb = openpyxl.load_workbook(xl_path)
        ws = wb["SolidWorks"]

        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1).value = self.proj_num
        ws.cell(row=next_row, column=2).value = self.cl_name
        ws.cell(row=next_row, column=3).value = self.dwg_type
        ws.cell(row=next_row, column=4).value = self.designer_name
        ws.cell(row=next_row, column=5).value = start_sw_time.strftime("%d/%m/%Y")
        ws.cell(row=next_row, column=6).value = self.current_sw_time.strftime("%H:%M:%S")
        ws.cell(row=next_row, column=7).value = stop_sw_time.strftime("%H:%M:%S")
        ws.cell(row=next_row, column=8).value = elapsed_sw_time
        ws.cell(row=next_row, column=9).value = self.remark

        wb.save(xl_path)

    # start GstarCAD multi-threading
    def start_gcad(self):
        self.logo_gcad_btn.config(state='disabled')

        t = threading.Thread(target=self.start_gcad_thread)
        t.start()

    # get User input, start GCAD & timer
    def start_gcad_thread(self):
        self.proj_num = self.projnum_entry.get()
        self.cl_name = self.clname_entry.get().title()
        self.dwg_type = self.dwg_type = self.dwg_type_cb.get()
        self.remark = self.note_text.get("1.0", 'end').capitalize()

        self.current_date = datetime.date.today()
        self.current_gcad_time = datetime.datetime.now()

        # Check users GCAD path that exists
        self.gcad_path_chk = os.path.exists(r"C:\Program Files\Gstarsoft\GstarCAD2022\gcad.exe")

        # for different user devices path
        if self.gcad_path_chk == True:
            subprocess.Popen(r'"C:\Program Files\Gstarsoft\GstarCAD2022\gcad.exe"')

        else:
            subprocess.Popen(r'"C:\Program Files\Gstarsoft\GstarCAD2023\gcad.exe"')

        self.timer_gcad_start()

        # check applications in the background
        while True:
            if any("gcad.exe" in p.name() for p in psutil.process_iter()):
                # Solidworks is running, continue loop
                continue

            else:
                self.stop_gcad()

                break

    # stop GCAD, open excel & save all info
    def stop_gcad(self):

        self.one_drive_path = os.environ.get('OneDriveCommercial') # get  OneDrive path
        xl_path_chk = [
            self.one_drive_path + r'\xxx\xxx\xxx\Project-Time-Table-2024.xlsx',
            self.one_drive_path + r'\xxx\xxx\xxx\xxx\Project-Time-Table-2024.xlsx',
            self.one_drive_path + r'\xxx\xxx\xxx\xxx\Project-Time-Table-2024.xlsx'
        ]

        self.designer = os.getlogin()
        self.designer_list = {
            'Device Name 1': 'Name 1',
            'Device Name 2': 'Name 2',
            'Device Name 3': 'Name 3',
            'Device Name 4': 'Name 4'
        }

        self.timer_gcad_stop()
        self.logo_gcad_btn.config(state='normal')

        stop_gcad_time = datetime.datetime.now()
        elapsed_gcad_time = stop_gcad_time - start_gcad_time

        # check where user excel path is
        for path in xl_path_chk:
            xlp = os.path.exists(path)

            if xlp == True:
                xl_path = path

        # check which user is signing in
        if self.designer in self.designer_list:
            self.designer_name = self.designer_list[self.designer]

        # open excel, insert all info, save then exit excel
        wb = openpyxl.load_workbook(xl_path)
        ws = wb["GStarCAD"]

        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1).value = self.proj_num
        ws.cell(row=next_row, column=2).value = self.cl_name
        ws.cell(row=next_row, column=3).value = self.dwg_type
        ws.cell(row=next_row, column=4).value = self.designer_name
        ws.cell(row=next_row, column=5).value = start_gcad_time.strftime("%d/%m/%Y")
        ws.cell(row=next_row, column=6).value = self.current_gcad_time.strftime("%H:%M:%S")
        ws.cell(row=next_row, column=7).value = stop_gcad_time.strftime("%H:%M:%S")
        ws.cell(row=next_row, column=8).value = elapsed_gcad_time
        ws.cell(row=next_row, column=9).value = self.remark

        wb.save(xl_path)

    # when tap the 'X' button, prompt the user
    def on_closing(self):
        if messagebox.askyesno(title="Quit", message="Have you save all your works?\nDid you close SolidWorks/GstarCAD?"):
            self.master.destroy()

root = tk.Tk()
root.geometry('360x400')
root.configure(bg='linen')
root.iconbitmap("PTT.ico")
swtt = swTimeTracker(root)

root.mainloop()
